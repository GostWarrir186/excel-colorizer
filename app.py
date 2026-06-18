from flask import Flask, request, send_file, jsonify
import openpyxl
from openpyxl.styles import PatternFill
import urllib.request
import json
import io

app = Flask(__name__)

RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

resume_url_store = {}

def get_usd_tjs_rate():
    url = 'https://open.er-api.com/v6/latest/USD'
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=10) as response:
        data = json.loads(response.read())
    return data['rates']['TJS']

@app.route('/colorize', methods=['POST'])
def colorize():
    rate = get_usd_tjs_rate()
    limit = 200 * rate


@app.route('/save-url', methods=['POST'])
def save_url():
    data = request.json
    resume_url_store['url'] = data['url']
    return jsonify({'ok': True})

@app.route('/get-url')
def get_url():
    return jsonify({'url': resume_url_store.get('url', '')})

    file = request.files['file']
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    

    header_row = [cell.value for cell in ws[1]]
    color_col = None
    sum_col = None
    name_col = None
    phone_col = None

    for i, h in enumerate(header_row):
        if h == '_rowColor': color_col = i + 1
        if h == 'Сумма': sum_col = i + 1
        if h == 'Имя получателя физ. лица': name_col = i + 1
        if h == 'Контактный номер': phone_col = i + 1

    # Группируем по получателю и считаем суммы
    groups = {}
    rows_data = []
    for row in ws.iter_rows(min_row=2):
        name = row[name_col - 1].value if name_col else ''
        phone = row[phone_col - 1].value if phone_col else ''
        key = str(name) + str(phone)
        amount = row[sum_col - 1].value if sum_col else 0
        color = row[color_col - 1].value if color_col else 'none'
        if key not in groups:
            groups[key] = 0
        groups[key] += (amount or 0)
        rows_data.append({'row': row, 'key': key, 'color': color})

    # Красим строки
    for item in rows_data:
        row = item['row']
        color = item['color']
        key = item['key']

        if color == 'red':
            for cell in row:
                if color_col and cell.column != color_col:
                    cell.fill = RED_FILL
        elif groups[key] > limit and color != 'red':
            for cell in row:
                if color_col and cell.column != color_col:
                    cell.fill = YELLOW_FILL

    # Удаляем колонку _rowColor
    if color_col:
        ws.delete_cols(color_col)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='reestr_checked.xlsx'
    )

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)
